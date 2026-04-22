"""
厚度分類功能測試
執行：pytest test_thickness_classify.py -v
"""
import pytest
import ezdxf
import openpyxl
from main import (
    build_model_with_properties,
    classify_by_thickness,
    classify_centerlines,
    classify_centerlines_full,
    classify_centerlines_from_geometry,
    classify_centerlines_from_geometry_full,
    extract_centerlines_with_thickness,
    write_dxf_classified,
    write_analytical_xlsx,
)


class TestClassifyByThickness:
    def test_thin_member_is_platform(self):
        assert classify_by_thickness(0.1) == "月台"

    def test_thick_member_is_main_structure(self):
        """三角測量：確認厚 0.5m 的構件確實回傳「主結構」，而非兩者都回傳「月台」。"""
        assert classify_by_thickness(0.5) == "主結構"

    def test_boundary_exactly_threshold_is_platform(self):
        """邊界：恰好 0.2m（= threshold）→ 應為「月台」（<= 包含等於）。"""
        assert classify_by_thickness(0.2) == "月台"

    def test_just_over_threshold_is_main_structure(self):
        """邊界：0.201m（略超 threshold）→ 應為「主結構」。"""
        assert classify_by_thickness(0.201) == "主結構"


class TestClassifyCenterlines:
    def test_single_thin_centerline(self):
        """一條薄中心線（0.1m）→ 回傳 [(中心線, '月台')]。"""
        cl = ((0.0, 0.0), (1.0, 0.0))
        result = classify_centerlines([(cl, 0.1)])
        assert result == [(cl, "月台")]

    def test_mixed_centerlines_classified_independently(self):
        """三角測量：混合輸入，每條依自己的厚度獨立分類，順序保留。"""
        cl1 = ((0.0, 0.0), (1.0, 0.0))  # 0.1m → 月台
        cl2 = ((0.0, 1.0), (1.0, 1.0))  # 0.5m → 主結構
        result = classify_centerlines([(cl1, 0.1), (cl2, 0.5)])
        assert result == [(cl1, "月台"), (cl2, "主結構")]


class TestClassifyCenterlinesFromGeometry:
    def test_mixed_geometry_produces_both_labels(self):
        """端對端：外框 2×1 + 內室（薄板 0.1m，厚牆 0.5m）→ 同時出現「月台」與「主結構」。
        幾何：outer 2×1、inner 0.5..1.5 × 0.1..0.9
          底板/頂板厚 0.1m ≤ 0.2m → 月台
          左牆/右牆厚 0.5m > 0.2m → 主結構"""
        outer = [(0.0, 0.0), (2.0, 0.0), (2.0, 1.0), (0.0, 1.0)]
        inner = [(0.5, 0.1), (1.5, 0.1), (1.5, 0.9), (0.5, 0.9)]
        labeled = classify_centerlines_from_geometry(outer, [inner], max_thickness=1.0)
        labels = {label for _, label in labeled}
        assert "月台" in labels
        assert "主結構" in labels


class TestClassifyCenterlinesFullWithThicknessPreserved:
    def test_returns_label_and_thickness_as_triple(self):
        """classify_centerlines_full 保留厚度：回傳 [(中心線, 標籤, 厚度), ...]。"""
        cl = ((0.0, 0.0), (1.0, 0.0))
        result = classify_centerlines_full([(cl, 0.15)])
        centerline, label, thickness = result[0]
        assert centerline == cl
        assert label == "月台"
        assert thickness == pytest.approx(0.15)

    def test_thick_member_preserves_main_structure_label_and_thickness(self):
        """三角測量：厚 0.5m 的構件 → 標籤「主結構」且厚度值保留為 0.5。"""
        cl = ((0.0, 0.0), (1.0, 0.0))
        result = classify_centerlines_full([(cl, 0.5)])
        centerline, label, thickness = result[0]
        assert centerline == cl
        assert label == "主結構"
        assert thickness == pytest.approx(0.5)


class TestClassifyCenterlinesFromGeometryFull:
    def test_geometry_returns_triples_with_both_labels(self):
        """幾何管線全輸出：outer 2×1 + inner（薄板 0.1m、厚牆 0.5m）→
        三元組同時出現「月台」與「主結構」，且各自厚度值正確保留。"""
        outer = [(0.0, 0.0), (2.0, 0.0), (2.0, 1.0), (0.0, 1.0)]
        inner = [(0.5, 0.1), (1.5, 0.1), (1.5, 0.9), (0.5, 0.9)]
        result = classify_centerlines_from_geometry_full(outer, [inner], max_thickness=1.0)
        labels = {label for _, label, _ in result}
        thicknesses_by_label = {label: t for _, label, t in result}
        assert "月台" in labels
        assert "主結構" in labels
        assert thicknesses_by_label["月台"] == pytest.approx(0.1)
        assert thicknesses_by_label["主結構"] == pytest.approx(0.5)


class TestWriteDxfClassified:
    def test_creates_platform_and_main_structure_layers(self, tmp_path):
        """分類 DXF 輸出：月台 → layer='PLATFORM'，主結構 → layer='MAIN_STRUCTURE'。"""
        cl1 = ((0.0, 0.0), (1.0, 0.0))
        cl2 = ((0.0, 1.0), (1.0, 1.0))
        labeled = [(cl1, "月台"), (cl2, "主結構")]
        out = str(tmp_path / "classified.dxf")
        write_dxf_classified(labeled, out)
        doc = ezdxf.readfile(out)
        layer_names = {e.dxf.layer for e in doc.modelspace()}
        assert "PLATFORM" in layer_names
        assert "MAIN_STRUCTURE" in layer_names



class TestBuildModelWithProperties:
    def test_elements_carry_label_and_thickness(self):
        """build_model_with_properties：兩條不相交的中心線 → 2 個桿件，
        各自帶有正確的 label 和 thickness。
        桿件格式：(id, n1, n2, label, thickness)。
        主結構先編號（ID=1），月台後編號（ID=2）。"""
        cl1 = ((0.0, 0.0), (4.0, 0.0))  # 水平，月台 0.1m
        cl2 = ((0.0, 2.0), (4.0, 2.0))  # 水平，主結構 0.5m
        triples = [(cl1, "月台", 0.1), (cl2, "主結構", 0.5)]
        _, elements = build_model_with_properties(triples)
        assert len(elements) == 2
        props = {eid: (label, t) for eid, _, _, label, t in elements}
        assert props[1] == ("主結構", pytest.approx(0.5))
        assert props[2] == ("月台", pytest.approx(0.1))

    def test_split_segments_inherit_source_thickness(self):
        """三角測量：兩條中心線相交 → split 後產生 4 個桿件，
        水平線的兩段均帶厚度 0.1m，垂直線的兩段均帶厚度 0.5m。"""
        cl_h = ((0.0, 1.0), (4.0, 1.0))  # 水平，月台 0.1m
        cl_v = ((2.0, 0.0), (2.0, 3.0))  # 垂直，主結構 0.5m
        triples = [(cl_h, "月台", 0.1), (cl_v, "主結構", 0.5)]
        _, elements = build_model_with_properties(triples)
        assert len(elements) == 4
        labels = [label for _, _, _, label, _ in elements]
        assert labels.count("月台") == 2
        assert labels.count("主結構") == 2
        h_thicknesses = [t for _, _, _, lbl, t in elements if lbl == "月台"]
        v_thicknesses = [t for _, _, _, lbl, t in elements if lbl == "主結構"]
        assert all(t == pytest.approx(0.1) for t in h_thicknesses)
        assert all(t == pytest.approx(0.5) for t in v_thicknesses)

    def test_main_structure_ids_precede_platform_ids(self):
        """排序：月台輸入在前，主結構在後 → 輸出中所有主結構 ID < 所有月台 ID。
        確保不管輸入順序，主結構桿件編號永遠先於月台桿件。"""
        cl1 = ((0.0, 0.0), (4.0, 0.0))  # 月台 — 輸入在前
        cl2 = ((0.0, 2.0), (4.0, 2.0))  # 主結構 — 輸入在後
        triples = [(cl1, "月台", 0.1), (cl2, "主結構", 0.5)]
        _, elements = build_model_with_properties(triples)
        main_ids = [eid for eid, _, _, lbl, _ in elements if lbl == "主結構"]
        platform_ids = [eid for eid, _, _, lbl, _ in elements if lbl == "月台"]
        assert max(main_ids) < min(platform_ids)

    def test_platform_globally_last_across_orientations(self):
        """三角測量：月台（水平板）的 ID 必須大於主結構（垂直柱）的 ID，
        即使月台板空間位置比主結構柱更低（確保「月台全域最後」跨越 H/V 分組）。
        若僅在各分組內部排月台最後（horizontals + verticals），月台-H 仍會排在主結構-V 之前。"""
        cl_main_v = ((2.0, 0.0), (2.0, 4.0))
        cl_platform_h = ((0.0, 1.0), (4.0, 1.0))
        triples = [(cl_main_v, "主結構", 0.5), (cl_platform_h, "月台", 0.1)]
        _, elements = build_model_with_properties(triples)
        main_ids = [eid for eid, _, _, lbl, _ in elements if lbl == "主結構"]
        platform_ids = [eid for eid, _, _, lbl, _ in elements if lbl == "月台"]
        assert max(main_ids) < min(platform_ids)


class TestSpatialOrdering:
    def test_nodes_ordered_bottom_row_first_then_left_to_right(self):
        """節點按「排」編號：同一 y 排由左到右（x 遞增），排與排之間由下往上（y 遞增）。
        底部線 y=0、頂部線 y=2，各在 x=0 和 x=4 有端點 → 4 個節點：
        底排左→右：ID=1:(0,0), ID=2:(4,0)；頂排左→右：ID=3:(0,2), ID=4:(4,2)。"""
        cl_bot = ((0.0, 0.0), (4.0, 0.0))
        cl_top = ((0.0, 2.0), (4.0, 2.0))
        triples = [(cl_bot, "主結構", 0.5), (cl_top, "主結構", 0.5)]
        nodes, _ = build_model_with_properties(triples)
        pos = {nid: (x, y) for nid, x, y in nodes}
        assert pos[1] == pytest.approx((0.0, 0.0))
        assert pos[2] == pytest.approx((4.0, 0.0))
        assert pos[3] == pytest.approx((0.0, 2.0))
        assert pos[4] == pytest.approx((4.0, 2.0))

    def test_elements_ordered_bottom_row_first_regardless_of_input_order(self):
        """三角測量：桿件依中點按（y 遞增, x 遞增）排序。
        輸入故意反序（頂部線在前、底部線在後）→ 輸出仍底部桿件 ID=1、頂部桿件 ID=2。"""
        cl_top = ((0.0, 2.0), (4.0, 2.0))  # 頂部 — 輸入在前（故意反序）
        cl_bot = ((0.0, 0.0), (4.0, 0.0))  # 底部 — 輸入在後
        triples = [(cl_top, "主結構", 0.5), (cl_bot, "主結構", 0.5)]
        nodes, elements = build_model_with_properties(triples)
        node_pos = {nid: (x, y) for nid, x, y in nodes}
        mid_y = {eid: (node_pos[n1][1] + node_pos[n2][1]) / 2
                 for eid, n1, n2, *_ in elements}
        assert mid_y[1] == pytest.approx(0.0)
        assert mid_y[2] == pytest.approx(2.0)

    def test_vertical_columns_numbered_left_column_first(self):
        """垂直柱按「欄」編號：最左欄從底到頂編完，再往右欄（與水平板相反）。
        左欄上段 mid_x=0, mid_y=3 vs 右欄下段 mid_x=4, mid_y=1：
        左欄段仍先編號（ID=1），右欄段後編號（ID=2）。"""
        cl_left_top = ((0.0, 2.0), (0.0, 4.0))   # 左欄上段，輸入在前
        cl_right_bot = ((4.0, 0.0), (4.0, 2.0))   # 右欄下段，輸入在後
        triples = [(cl_left_top, "主結構", 0.5), (cl_right_bot, "主結構", 0.5)]
        nodes, elements = build_model_with_properties(triples)
        node_pos = {nid: (x, y) for nid, x, y in nodes}
        mid_x = {eid: (node_pos[n1][0] + node_pos[n2][0]) / 2
                 for eid, n1, n2, *_ in elements}
        assert mid_x[1] == pytest.approx(0.0)  # 左欄先
        assert mid_x[2] == pytest.approx(4.0)  # 右欄後

    def test_horizontal_plates_numbered_before_vertical_columns(self):
        """水平板先於垂直柱編號，不論空間位置或輸入順序。
        垂直柱中點 y=2（低）、水平板中點 y=5（高）→ 板仍得 ID=1，柱得 ID=2。"""
        cl_v = ((2.0, 0.0), (2.0, 4.0))   # 垂直柱，mid_y=2，輸入在前
        cl_h = ((0.0, 5.0), (4.0, 5.0))   # 水平板，mid_y=5，輸入在後
        triples = [(cl_v, "主結構", 0.5), (cl_h, "主結構", 0.5)]
        nodes, elements = build_model_with_properties(triples)
        node_pos = {nid: (x, y) for nid, x, y in nodes}
        is_horiz = {eid: abs(node_pos[n1][1] - node_pos[n2][1]) < 1e-3
                    for eid, n1, n2, *_ in elements}
        assert is_horiz[1] is True,  "ID=1 應為水平板"
        assert is_horiz[2] is False, "ID=2 應為垂直柱"


class TestExtractCenterlinesWithThickness:
    def test_geometry_produces_thickness_pairs(self):
        """整合：外框 6×4 + 內室各邊 1m 厚，max_thickness=2.0 → 4 條主構件厚度均為 1.0m。
        不加 max_thickness 則角落殘餘段會與外框對側配對得厚度 4.0m（此為正確演算法行為，
        測試此處明確給定上限以隔離主構件）。"""
        outer = [(0.0, 0.0), (6.0, 0.0), (6.0, 4.0), (0.0, 4.0)]
        inner = [(1.0, 1.0), (5.0, 1.0), (5.0, 3.0), (1.0, 3.0)]
        pairs = extract_centerlines_with_thickness(outer, [inner], max_thickness=2.0)
        thicknesses = [t for _, t in pairs]
        assert all(t == pytest.approx(1.0) for t in thicknesses)


class TestWriteAnalyticalXlsx:
    def test_creates_three_sheets_with_correct_names(self, tmp_path):
        """write_analytical_xlsx 建立含 3 個 sheet 的 xlsx：
        依序為 'C.結構點位'、'D.桿件編號'、'E.桿件資訊'。"""
        nodes = [(1, 0.0, 0.0), (2, 4.0, 0.0)]
        elements = [(1, 1, 2, "主結構", 0.5)]
        out = str(tmp_path / "test.xlsx")
        write_analytical_xlsx(nodes, elements, out)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["C.結構點位", "D.桿件編號", "E.桿件資訊"]

    def test_node_coordinates_normalized_to_min_xy_as_origin(self, tmp_path):
        """C.結構點位：以最左下角 (min_x, min_y) 為原點 (0,0) 輸出座標。
        節點 (2.0, 3.0) 和 (5.0, 7.0) → 第 4 列起輸出 (0.0, 0.0) 和 (3.0, 4.0)。"""
        nodes = [(1, 2.0, 3.0), (2, 5.0, 7.0)]
        elements = [(1, 1, 2, "主結構", 0.5)]
        out = str(tmp_path / "test.xlsx")
        write_analytical_xlsx(nodes, elements, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["C.結構點位"]
        data_rows = [row for row in ws.iter_rows(min_row=4, values_only=True)
                     if row[0] is not None and isinstance(row[0], int)]
        assert data_rows[0][1] == pytest.approx(0.0)  # node 1 x
        assert data_rows[0][2] == pytest.approx(0.0)  # node 1 y
        assert data_rows[1][1] == pytest.approx(3.0)  # node 2 x
        assert data_rows[1][2] == pytest.approx(4.0)  # node 2 y

    def test_element_connectivity_in_sheet_d(self, tmp_path):
        """D.桿件編號：每桿件輸出 [桿件編號, 點位1, 點位2]，順序與 elements 一致。"""
        nodes = [(1, 0.0, 0.0), (2, 4.0, 0.0), (3, 0.0, 3.0)]
        elements = [(1, 1, 2, "主結構", 0.5), (2, 1, 3, "主結構", 0.5)]
        out = str(tmp_path / "test.xlsx")
        write_analytical_xlsx(nodes, elements, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["D.桿件編號"]
        data_rows = [row for row in ws.iter_rows(min_row=4, values_only=True)
                     if row[0] is not None and isinstance(row[0], int)]
        assert data_rows[0][:3] == (1, 1, 2)
        assert data_rows[1][:3] == (2, 1, 3)

    def test_thickness_in_sheet_e_width_column(self, tmp_path):
        """E.桿件資訊：桿件厚度（元素第 5 欄 thickness）寫入第 2 欄（桿件寬(m)）。
        元素格式 (eid, n1, n2, label, thickness)，期待輸出 [eid, thickness]。"""
        nodes = [(1, 0.0, 0.0), (2, 4.0, 0.0)]
        elements = [(1, 1, 2, "主結構", 0.5)]
        out = str(tmp_path / "test.xlsx")
        write_analytical_xlsx(nodes, elements, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["E.桿件資訊"]
        data_rows = [row for row in ws.iter_rows(min_row=4, values_only=True)
                     if row[0] is not None and isinstance(row[0], int)]
        assert data_rows[0][0] == 1
        assert data_rows[0][1] == pytest.approx(0.5)
