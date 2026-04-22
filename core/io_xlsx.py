# Excel (xlsx) 輸出

import openpyxl


def write_analytical_xlsx(nodes, elements, filepath):
    """將所有資訊與節點（C.結構點位）、桿件（D.桿件編號）合併輸出為單一 xlsx。"""
    wb = openpyxl.Workbook()

    # --- Sheet 1: A.本案資訊 ---
    ws_a = wb.active
    ws_a.title = "A.本案資訊"
    for row in [
        ["A.本案資訊"], [],
        ["計畫："], ["子標："], ["日期："], ["設計者："], ["校核者："], ["斷面名稱："]
    ]:
        ws_a.append(row)

    # --- Sheet 2: B.基本資訊 ---
    ws_b = wb.create_sheet("B.基本資訊")
    for row in [
        ["B.基本資訊"], [],
        ["地表高程(m)："], ["洪水位高程(m)："], ["常水位高程(m)："], ["低水位高程(m)："],
        ["不對稱高水位高程(m)："], ["不對稱低水位高程(m)："], ["土壤液化高程(高)(m)："],
        ["土壤液化高程(低)(m)："], ["底版頂高程(m)："], ["結構剪力角："], ["垂直土壤彈性值(t/m3)："],
        ["土壤單位重(t/m3)："], ["混凝土單位重(t/m3)："], ["水單位重(t/m3)："], ["浸水土壤單位重(t/m3)："],
        ["混凝土彈性模數(tf/m2)："], ["鋼筋彈性模數(tf/m2)："], ["柏松比："], ["鋼筋之規定降伏強度(kgf/cm2)："],
        ["混凝土規定抗壓強度(kgf/cm2)："], ["動態土壓增量係數(ODE)："], ["動態土壓增量係數(MDE)："],
        ["靜態土壓力係數："], ["中度地震地表加速度(g)："], ["最大地震地表加速度(g)："], ["電聯車輪軸重(kN)："],
        ["電聯車軌距(m)："], ["電聯車車軸間距(m)："], ["輪載重平均分佈寬(m)："], ["輪載重平均分佈長(m)："],
        ["剖面厚度(m)："], ["起始里程："], ["結束里程："], ["水平土壤彈性值(t/m3)："], ["開挖地表高程(m)："],
        ["主動土壓力係數："]
    ]:
        ws_b.append(row)

    # --- Sheet 3: C.結構點位 ---
    ws_n = wb.create_sheet("C.結構點位")
    xs = [x for _, x, _ in nodes]
    ys = [y for _, _, y in nodes]
    ox, oy = (min(xs), min(ys)) if nodes else (0, 0)
    ws_n.append(["C.結構點位"])
    ws_n.append([])
    ws_n.append(["起始點位編號", "X(m)", "Y(m)", "Z(m)",
                 "終端點位編號", "X(m)", "Y(m)", "Z(m)"])
    for nid, x, y in nodes:
        ws_n.append([nid, round(x - ox, 3), round(y - oy, 3), 0])
    ws_n.append([])
    ws_n.append
    ws_n.append(["輸入方式:"])
    ws_n.append(["(1)單節點"])
    ws_n.append(["點位編號", "x", "y", "z"])
    ws_n.append(["(2)多節點相同間隔"])
    ws_n.append(["點位編號(起始)", "x1", "y1", "z1", "點位編號(終端)", "x2", "y2", "z2"])

    # --- Sheet 2: D.桿件編號 ---
    ws_e = wb.create_sheet("D.桿件編號")
    ws_e.append(["D.桿件編號"])
    ws_e.append([])
    ws_e.append(["桿件編號(起始)", "點位1", "點位2", "桿件編號(終端)"])
    for eid, n1, n2, *_ in elements:
        ws_e.append([eid, n1, n2])
    ws_e.append([])
    ws_e.append(["-----------------------------------------------------------------------"])
    ws_e.append(["輸入方式:"])
    ws_e.append(["(1)單桿件"])
    ws_e.append(["桿件編號", "點位1", "點位2"])
    ws_e.append(["(2)多桿件相同邏輯"])
    ws_e.append(["桿件編號(起始)", "點位1", "點位1+n", "桿件編號(終端)"])

    # --- Sheet 3: E.桿件資訊 ---
    ws_i = wb.create_sheet("E.桿件資訊")
    ws_i.append(["E.桿件資訊"])
    ws_i.append([])
    ws_i.append(["桿件編號(起始)", "桿件寬(m)", "剖面厚(m)", "桿件編號(終端)", "自重歸類到b case", "fc'"])
    for eid, n1, n2, *extra in elements:
        thickness = extra[1] if len(extra) >= 2 else None
        ws_i.append([eid, thickness])
    ws_i.append([])
    ws_i.append(["-----------------------------------------------------------------------"])
    ws_i.append(["輸入方式:"])
    ws_i.append(["(1)單桿件"])
    ws_i.append(["桿件編號", "點位1", "點位2"])
    ws_i.append(["(2)多桿件相同邏輯"])
    ws_i.append(["桿件編號(起始)", "點位1", "點位1+n", "桿件編號(終端)"])

    wb.save(filepath)
